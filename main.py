import os
import re
from datetime import datetime
from urllib.parse import quote

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# FONTES - NOTÍCIAS AGRÍCOLAS
# ============================================================

URL_MILHO = "https://www.noticiasagricolas.com.br/cotacoes/milho/milho-b3-prego-regular"
URL_SOJA = "https://www.noticiasagricolas.com.br/cotacoes/soja/soja-bolsa-de-chicago-cme-group"
URL_DOLAR = "https://www.noticiasagricolas.com.br/cotacao-do-dolar/"

# ============================================================
# FONTES - SIMA / CELEPAR
# ============================================================

URL_SIMA_INICIAL = "https://celepar7.pr.gov.br/sima/cotdiap.asp"
URL_SIMA_RESULTADO = "https://celepar7.pr.gov.br/sima/cotdiap1.asp"

PRODUTOS_SIMA = [
    {
        "codigo": "7",
        "grao": "MILHO",
        "produto": "Milho amarelo tipo 1 sc 60 Kg",
    },
    {
        "codigo": "8",
        "grao": "SOJA",
        "produto": "Soja industrial tipo 1 sc 60 Kg",
    },
]

# ============================================================
# EXCEL
# ============================================================

ARQUIVO_EXCEL = "cotacoes_graos.xlsx"

ABA_COTACOES = "cotacoes"
ABA_SIMA = "sima_pracas"

TABELA_COTACOES = "tb_cotacoes_graos"
TABELA_SIMA = "tb_sima_pracas"

CABECALHOS_COTACOES = [
    "data_coleta",
    "data_cotacao",
    "grao",
    "contrato_mes",
    "fechamento_rs_sc_60kg",
    "fechamento_usd_bushel",
    "dolar",
    "fonte",
    "url",
    "chave",
]

CABECALHOS_SIMA = [
    "data_coleta",
    "data_cotacao",
    "grao",
    "produto",
    "nucleo_regional",
    "m_c",
    "fonte",
    "url",
    "chave",
]


# ============================================================
# FUNÇÕES GERAIS
# ============================================================

def numero_br_para_float(valor):
    if valor is None:
        return None

    valor = str(valor).strip()

    if valor.upper() in ["SINF", "AUS", "-", ""]:
        return None

    if not re.search(r"\d", valor):
        return None

    if any(palavra in valor.upper() for palavra in ["PREÇO", "PRECO", "COMUM", "M_C", "MIN", "MAX"]):
        return None

    valor = valor.replace("+", "")
    valor = valor.replace(".", "")
    valor = valor.replace(",", ".")

    try:
        return float(valor)
    except ValueError:
        return None


def baixar_texto_pagina(url):
    headers = {
        "User-Agent": "Mozilla/5.0"
    }

    print(f"Acessando: {url}")

    response = requests.get(url, headers=headers, timeout=30)
    response.encoding = response.apparent_encoding
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    texto = soup.get_text("\n", strip=True)

    return texto


def data_hora_coleta():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ============================================================
# NOTÍCIAS AGRÍCOLAS - DÓLAR
# ============================================================

def buscar_dolar():
    texto = baixar_texto_pagina(URL_DOLAR)

    # Procura dólar no formato R$ 5,00 / R$ 5,67 / R$ 5.67
    match = re.search(r"R\$\s*(\d+[,.]\d{2,4})", texto)

    if not match:
        print("Não consegui encontrar o dólar. Texto parcial da página:")
        print(texto[:1500])
        raise Exception("Dólar não encontrado.")

    dolar = numero_br_para_float(match.group(1))
    dolar = round(dolar, 2)

    print(f"Dólar encontrado: {dolar:.2f}")

    return dolar


# ============================================================
# NOTÍCIAS AGRÍCOLAS - MILHO B3 / SOJA CHICAGO
# ============================================================

def extrair_bloco_mais_recente(texto):
    match_data = re.search(r"Fechamento:\s*(\d{2}/\d{2}/\d{4})", texto)

    if not match_data:
        print("Texto parcial da página:")
        print(texto[:2000])
        raise Exception("Data de fechamento não encontrada.")

    data_cotacao = match_data.group(1)

    inicio = match_data.end()

    proximo_fechamento = re.search(r"Fechamento:\s*\d{2}/\d{2}/\d{4}", texto[inicio:])

    if proximo_fechamento:
        fim = inicio + proximo_fechamento.start()
        bloco = texto[inicio:fim]
    else:
        bloco = texto[inicio:]

    return data_cotacao, bloco


def buscar_cotacoes_milho_b3(dolar):
    texto = baixar_texto_pagina(URL_MILHO)

    data_cotacao, bloco = extrair_bloco_mais_recente(texto)

    linhas = []

    padrao = re.compile(
        r"([A-Za-zçÇãÃéÉíÍóÓúÚ]+/\d{4})\s+([\d,]+)\s+([-+]?[\d,]+)"
    )

    for match in padrao.finditer(bloco):
        contrato_mes = match.group(1)
        fechamento = numero_br_para_float(match.group(2))

        linha = {
            "data_coleta": data_hora_coleta(),
            "data_cotacao": data_cotacao,
            "grao": "MILHO",
            "contrato_mes": contrato_mes,
            "fechamento_rs_sc_60kg": fechamento,
            "fechamento_usd_bushel": None,
            "dolar": dolar,
            "fonte": "Notícias Agrícolas / B3",
            "url": URL_MILHO,
            "chave": f"MILHO_B3_{contrato_mes}_{data_cotacao}",
        }

        linhas.append(linha)

    if not linhas:
        print("AVISO: Nenhuma cotação de milho B3 encontrada neste momento.")
        print("Bloco de milho retornado pelo site:")
        print(bloco[:2000])
        return []

    print(f"Cotações de milho B3 encontradas: {len(linhas)}")

    return linhas


def buscar_cotacoes_soja_chicago(dolar):
    texto = baixar_texto_pagina(URL_SOJA)

    data_cotacao, bloco = extrair_bloco_mais_recente(texto)

    linhas = []

    padrao = re.compile(
        r"([A-Za-zçÇãÃéÉíÍóÓúÚ]+/\d{2})\s+([\d,]+)\s+([-+]?[\d,]+)\s+([-+]?[\d,]+)"
    )

    for match in padrao.finditer(bloco):
        contrato_mes = match.group(1)
        fechamento_usd_bushel = numero_br_para_float(match.group(2))

        linha = {
            "data_coleta": data_hora_coleta(),
            "data_cotacao": data_cotacao,
            "grao": "SOJA",
            "contrato_mes": contrato_mes,
            "fechamento_rs_sc_60kg": None,
            "fechamento_usd_bushel": fechamento_usd_bushel,
            "dolar": dolar,
            "fonte": "Notícias Agrícolas / CME Group",
            "url": URL_SOJA,
            "chave": f"SOJA_CHICAGO_{contrato_mes}_{data_cotacao}",
        }

        linhas.append(linha)

    if not linhas:
        print("AVISO: Nenhuma cotação de soja Chicago encontrada neste momento.")
        print("Bloco de soja retornado pelo site:")
        print(bloco[:2000])
        return []

    print(f"Cotações de soja Chicago encontradas: {len(linhas)}")

    return linhas


# ============================================================
# SIMA / CELEPAR
# ============================================================

def baixar_resultado_sima(codigo_produto):
    session = requests.Session()

    headers_get = {
        "User-Agent": "Mozilla/5.0"
    }

    response_inicial = session.get(
        URL_SIMA_INICIAL,
        headers=headers_get,
        timeout=30
    )

    response_inicial.encoding = response_inicial.apparent_encoding
    response_inicial.raise_for_status()

    headers_post = {
        "User-Agent": "Mozilla/5.0",
        "Referer": URL_SIMA_INICIAL,
        "Content-Type": "application/x-www-form-urlencoded",
    }

    dados = {
        "produto": codigo_produto,
        "submit1": "Pesquisar",
    }

    response = session.post(
        URL_SIMA_RESULTADO,
        data=dados,
        headers=headers_post,
        timeout=30
    )

    response.encoding = response.apparent_encoding
    response.raise_for_status()

    return response.text


def extrair_data_sima(texto):
    match = re.search(r"em\s+(\d{2}/\d{2}/\d{4})", texto)

    if match:
        return match.group(1)

    return None


def extrair_linhas_sima_m_c(html, grao, produto_nome):
    soup = BeautifulSoup(html, "html.parser")
    texto = soup.get_text("\n", strip=True)

    data_cotacao = extrair_data_sima(texto)

    print(f"Data SIMA encontrada para {grao}: {data_cotacao}")

    tabelas = soup.find_all("table")

    if not tabelas:
        print(f"AVISO: Nenhuma tabela SIMA encontrada para {produto_nome}.")
        print(texto[:2000])
        return []

    linhas_extraidas = []

    for tabela in tabelas:
        linhas = tabela.find_all("tr")

        for linha in linhas:
            colunas = linha.find_all(["td", "th"])
            valores = [col.get_text(strip=True) for col in colunas]

            if len(valores) < 4:
                continue

            nucleo_regional = valores[0].strip()
            minimo = valores[1].strip()
            m_c = valores[2].strip()
            maximo = valores[3].strip()

            texto_linha = " ".join(valores).upper()

            # Ignora cabeçalho
            if "NÚCLEO REGIONAL" in texto_linha or "NUCLEO REGIONAL" in texto_linha:
                continue

            # Ignora legenda/rodapé
            if "PREÇO" in texto_linha or "PRECO" in texto_linha:
                continue

            if "FONTE" in texto_linha:
                continue

            m_c_numero = numero_br_para_float(m_c)

            if m_c_numero is None:
                continue

            linha_saida = {
                "data_coleta": data_hora_coleta(),
                "data_cotacao": data_cotacao,
                "grao": grao,
                "produto": produto_nome,
                "nucleo_regional": nucleo_regional,
                "m_c": m_c_numero,
                "fonte": "SIMA / CELEPAR / SEAB-PR",
                "url": URL_SIMA_INICIAL,
                "chave": f"SIMA_{grao}_{nucleo_regional}_{data_cotacao}",
            }

            linhas_extraidas.append(linha_saida)

    return linhas_extraidas


def buscar_cotacoes_sima():
    todas_linhas = []

    for produto in PRODUTOS_SIMA:
        print("=" * 80)
        print(f"Coletando SIMA: {produto['produto']}")
        print(f"Código interno: {produto['codigo']}")

        html = baixar_resultado_sima(produto["codigo"])

        linhas = extrair_linhas_sima_m_c(
            html=html,
            grao=produto["grao"],
            produto_nome=produto["produto"],
        )

        print(f"Linhas SIMA encontradas para {produto['grao']}: {len(linhas)}")

        todas_linhas.extend(linhas)

    print(f"Total de linhas SIMA encontradas: {len(todas_linhas)}")

    return todas_linhas


# ============================================================
# EXCEL - CRIAR / ABRIR
# ============================================================

def criar_ou_abrir_workbook():
    if os.path.exists(ARQUIVO_EXCEL):
        print("Arquivo Excel já existe. Abrindo...")
        workbook = load_workbook(ARQUIVO_EXCEL)
    else:
        print("Arquivo Excel não existe. Criando...")
        workbook = Workbook()

        # Remove aba padrão se estiver vazia
        aba_padrao = workbook.active
        workbook.remove(aba_padrao)

    return workbook


def obter_ou_criar_aba(workbook, nome_aba, cabecalhos):
    if nome_aba in workbook.sheetnames:
        sheet = workbook[nome_aba]
    else:
        print(f"Criando aba: {nome_aba}")
        sheet = workbook.create_sheet(nome_aba)

    # Se a aba estiver vazia, cria cabeçalho
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
        sheet.append(cabecalhos)

    # Se a primeira linha não tiver os cabeçalhos esperados, garante o cabeçalho
    primeira_linha = [sheet.cell(row=1, column=i + 1).value for i in range(len(cabecalhos))]

    if primeira_linha != cabecalhos:
        for i, cabecalho in enumerate(cabecalhos, start=1):
            sheet.cell(row=1, column=i).value = cabecalho

    return sheet


def carregar_chaves_existentes(sheet, indice_coluna_chave):
    chaves = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) <= indice_coluna_chave:
            continue

        chave = row[indice_coluna_chave]

        if chave and chave != "chave":
            chaves.add(chave)

    return chaves


# ============================================================
# EXCEL - FORMATAÇÃO E TABELAS
# ============================================================

def formatar_aba_cotacoes(sheet):
    for cell in sheet["A"]:
        if cell.row == 1:
            continue
        cell.number_format = "yyyy-mm-dd hh:mm:ss"

    for cell in sheet["E"]:
        if cell.row == 1:
            continue
        cell.number_format = "0.00"

    for cell in sheet["F"]:
        if cell.row == 1:
            continue
        cell.number_format = "0.0000"

    for cell in sheet["G"]:
        if cell.row == 1:
            continue
        cell.number_format = "0.00"

    larguras = {
        "A": 22,
        "B": 15,
        "C": 12,
        "D": 18,
        "E": 24,
        "F": 24,
        "G": 12,
        "H": 32,
        "I": 80,
        "J": 40,
    }

    for coluna, largura in larguras.items():
        sheet.column_dimensions[coluna].width = largura


def formatar_aba_sima(sheet):
    for cell in sheet["A"]:
        if cell.row == 1:
            continue
        cell.number_format = "yyyy-mm-dd hh:mm:ss"

    for cell in sheet["F"]:
        if cell.row == 1:
            continue
        cell.number_format = "0.00"

    larguras = {
        "A": 22,
        "B": 15,
        "C": 12,
        "D": 34,
        "E": 26,
        "F": 12,
        "G": 32,
        "H": 65,
        "I": 45,
    }

    for coluna, largura in larguras.items():
        sheet.column_dimensions[coluna].width = largura


def criar_ou_atualizar_tabela_excel(sheet, nome_tabela, quantidade_colunas):
    ultima_linha = sheet.max_row

    if ultima_linha < 2:
        print(f"Não há linhas suficientes para criar tabela na aba {sheet.title}.")
        return

    ultima_coluna_letra = get_column_letter(quantidade_colunas)
    intervalo_tabela = f"A1:{ultima_coluna_letra}{ultima_linha}"

    if nome_tabela in sheet.tables:
        print(f"Tabela {nome_tabela} já existe. Atualizando intervalo para {intervalo_tabela}...")
        sheet.tables[nome_tabela].ref = intervalo_tabela
    else:
        print(f"Criando tabela {nome_tabela} no intervalo {intervalo_tabela}...")

        tabela = Table(
            displayName=nome_tabela,
            ref=intervalo_tabela
        )

        estilo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        tabela.tableStyleInfo = estilo
        sheet.add_table(tabela)

    sheet.freeze_panes = "A2"


# ============================================================
# EXCEL - SALVAR DADOS
# ============================================================

def salvar_cotacoes_na_aba(sheet, linhas):
    chaves_existentes = carregar_chaves_existentes(sheet, indice_coluna_chave=9)

    novas_linhas = 0
    duplicadas = 0

    for cotacao in linhas:
        if cotacao["chave"] in chaves_existentes:
            print(f"Já existe em cotacoes, não vou duplicar: {cotacao['chave']}")
            duplicadas += 1
            continue

        nova_linha = [
            cotacao["data_coleta"],
            cotacao["data_cotacao"],
            cotacao["grao"],
            cotacao["contrato_mes"],
            cotacao["fechamento_rs_sc_60kg"],
            cotacao["fechamento_usd_bushel"],
            cotacao["dolar"],
            cotacao["fonte"],
            cotacao["url"],
            cotacao["chave"],
        ]

        sheet.append(nova_linha)
        chaves_existentes.add(cotacao["chave"])
        novas_linhas += 1

    print(f"Aba cotacoes - novas linhas adicionadas: {novas_linhas}")
    print(f"Aba cotacoes - duplicadas ignoradas: {duplicadas}")


def salvar_sima_na_aba(sheet, linhas):
    chaves_existentes = carregar_chaves_existentes(sheet, indice_coluna_chave=8)

    novas_linhas = 0
    duplicadas = 0

    for cotacao in linhas:
        if cotacao["chave"] in chaves_existentes:
            print(f"Já existe em sima_pracas, não vou duplicar: {cotacao['chave']}")
            duplicadas += 1
            continue

        nova_linha = [
            cotacao["data_coleta"],
            cotacao["data_cotacao"],
            cotacao["grao"],
            cotacao["produto"],
            cotacao["nucleo_regional"],
            cotacao["m_c"],
            cotacao["fonte"],
            cotacao["url"],
            cotacao["chave"],
        ]

        sheet.append(nova_linha)
        chaves_existentes.add(cotacao["chave"])
        novas_linhas += 1

    print(f"Aba sima_pracas - novas linhas adicionadas: {novas_linhas}")
    print(f"Aba sima_pracas - duplicadas ignoradas: {duplicadas}")


def salvar_no_excel(linhas_cotacoes, linhas_sima):
    workbook = criar_ou_abrir_workbook()

    sheet_cotacoes = obter_ou_criar_aba(
        workbook=workbook,
        nome_aba=ABA_COTACOES,
        cabecalhos=CABECALHOS_COTACOES,
    )

    sheet_sima = obter_ou_criar_aba(
        workbook=workbook,
        nome_aba=ABA_SIMA,
        cabecalhos=CABECALHOS_SIMA,
    )

    salvar_cotacoes_na_aba(sheet_cotacoes, linhas_cotacoes)
    salvar_sima_na_aba(sheet_sima, linhas_sima)

    formatar_aba_cotacoes(sheet_cotacoes)
    formatar_aba_sima(sheet_sima)

    criar_ou_atualizar_tabela_excel(
        sheet=sheet_cotacoes,
        nome_tabela=TABELA_COTACOES,
        quantidade_colunas=len(CABECALHOS_COTACOES),
    )

    criar_ou_atualizar_tabela_excel(
        sheet=sheet_sima,
        nome_tabela=TABELA_SIMA,
        quantidade_colunas=len(CABECALHOS_SIMA),
    )

    workbook.save(ARQUIVO_EXCEL)

    print(f"Arquivo salvo: {ARQUIVO_EXCEL}")


# ============================================================
# MICROSOFT GRAPH / SHAREPOINT
# ============================================================

def obter_token_graph():
    tenant_id = os.environ["TENANT_ID"]
    client_id = os.environ["CLIENT_ID"]
    client_secret = os.environ["CLIENT_SECRET"]

    url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"

    payload = {
        "client_id": client_id,
        "client_secret": client_secret,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }

    response = requests.post(url, data=payload, timeout=30)
    response.raise_for_status()

    return response.json()["access_token"]


def obter_site_id(token):
    hostname = os.environ["SHAREPOINT_HOSTNAME"]
    site_path = os.environ["SHAREPOINT_SITE_PATH"]

    headers = {
        "Authorization": f"Bearer {token}"
    }

    if site_path == "/":
        url = f"https://graph.microsoft.com/v1.0/sites/{hostname}"
    else:
        url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:{site_path}"

    response = requests.get(url, headers=headers, timeout=30)
    response.raise_for_status()

    site_id = response.json()["id"]

    print(f"Site ID encontrado: {site_id}")

    return site_id


def enviar_excel_para_sharepoint():
    print("Enviando Excel para o SharePoint...")

    token = obter_token_graph()
    site_id = obter_site_id(token)

    folder_path = os.environ["SHAREPOINT_FOLDER_PATH"].strip("/")
    arquivo_destino = ARQUIVO_EXCEL

    encoded_path = quote(f"{folder_path}/{arquivo_destino}", safe="/")

    url_upload = (
        f"https://graph.microsoft.com/v1.0/sites/{site_id}"
        f"/drive/root:/{encoded_path}:/content"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }

    with open(ARQUIVO_EXCEL, "rb") as arquivo:
        response = requests.put(url_upload, headers=headers, data=arquivo, timeout=60)

    response.raise_for_status()

    print("Arquivo enviado para o SharePoint com sucesso.")
    print(f"Destino: {folder_path}/{arquivo_destino}")


# ============================================================
# MAIN
# ============================================================

def main():
    dolar = buscar_dolar()

    linhas_milho_b3 = buscar_cotacoes_milho_b3(dolar)
    linhas_soja_chicago = buscar_cotacoes_soja_chicago(dolar)

    linhas_cotacoes = linhas_milho_b3 + linhas_soja_chicago

    print(f"Total de linhas Notícias Agrícolas: {len(linhas_cotacoes)}")

    linhas_sima = buscar_cotacoes_sima()

    print(f"Total de linhas SIMA/CELEPAR: {len(linhas_sima)}")

    if not linhas_cotacoes and not linhas_sima:
        raise Exception("Nenhuma cotação foi capturada. O fluxo será interrompido.")

    salvar_no_excel(
        linhas_cotacoes=linhas_cotacoes,
        linhas_sima=linhas_sima,
    )

    enviar_excel_para_sharepoint()


if __name__ == "__main__":
    main()
